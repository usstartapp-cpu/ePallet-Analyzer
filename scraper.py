"""
ePallet.com Product Scraper v2.0 — API-Based
=============================================
Uses the internal product search API for maximum speed and data quality.
Logs in via browser, then makes direct API calls for each category.

Usage:
  python3 scraper.py

Output: epallet_dry_products.xlsx
"""

import asyncio
import csv
import json
import math
import os
import re
import sys
import time
from datetime import datetime
from urllib.parse import unquote

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from playwright.async_api import async_playwright

from config import (
    BASE_URL,
    CHECKPOINT_CSV,
    EMAIL,
    HEADLESS,
    MAX_RETRIES,
    OUTPUT_EXCEL,
    PAGE_DELAY,
    PASSWORD,
    STORAGE_FILTER,
    TIMEOUT,
)


# ═══════════════════════════════════════════════════════════════════
# CONSTANTS
# ═══════════════════════════════════════════════════════════════════

# All Food categories discovered from /api/ERP/common/categories/common-categories
FOOD_CATEGORIES = [
    "Pantry & Baking Supplies",
    "Canned Goods",
    "Ingredients",
    "Meal Kits",
    "Condiments, Dressings, Spreads",
    "Baby Food",
    "Meat & Seafood",
    "Natural Foods & Plant-Based",
    "Fruits & Vegetables",
    "Frozen",
    "Eggs & Dairy",
    "Snacks",
    "Ready to Eat Meals",
    "Industrial Ingredients",
    "Beverages & Beverage Mixes",
    "Prepared & Ethnic Foods",
    "Baked Goods & Desserts",
    "Spices, Seasonings, Sweeteners",
    "Bakery, Tortillas",
    "Grains & Cereals",
    "Legumes, Nuts & Seeds",
]

# Non-Food categories (in case they want these too)
NONFOOD_CATEGORIES = [
    "Hand Sanitizers, Disinfectants & Wipes",
    "Disposable Serviceware",
    "Household",
    "Pet Products",
    "Medical Supplies",
    "Health & Beauty",
    "Baby Products",
    "Cleaning Supplies & Paper",
    "Janitorial Supplies",
    "PPE",
    "Pet Food",
    "Paper & Serviceware",
]

# API allows up to 200 items per page
API_PAGE_SIZE = 200

# Shipping zip code (from the account — Stafford, TX)
POST_ZIP = "77477"


# ═══════════════════════════════════════════════════════════════════
# AUTHENTICATION
# ═══════════════════════════════════════════════════════════════════

async def login(page):
    """Log in to ePallet via browser to establish session cookies."""
    print("\n" + "=" * 60)
    print("🔐 LOGGING IN TO EPALLET.COM")
    print("=" * 60)

    await page.goto(BASE_URL, wait_until="domcontentloaded")
    await page.wait_for_timeout(2000)

    # Dismiss cookies
    try:
        btn = page.locator("button:has-text('Accept All')")
        if await btn.is_visible(timeout=3000):
            await btn.click()
            print("  ✓ Cookie banner dismissed")
    except:
        pass

    # Click Sign In
    print("  → Clicking Sign In...")
    await page.locator("text=Sign In").first.click()
    await page.wait_for_timeout(2000)

    # Fill credentials
    print(f"  → Entering credentials ({EMAIL})...")
    await page.locator("input[placeholder*='email' i]").first.fill(EMAIL)
    await page.locator("input[type='password']").first.fill(PASSWORD)

    # Submit
    print("  → Submitting...")
    await page.locator("button[type='submit']").first.click()
    await page.wait_for_timeout(5000)

    # Verify
    status = await page.evaluate("""
        async () => {
            try {
                const resp = await fetch('/api/ERP/customer/status');
                return await resp.json();
            } catch(e) {
                return { error: e.message };
            }
        }
    """)

    if status.get("is_authenticated") or status.get("is_customer"):
        print(f"  ✅ LOGIN SUCCESSFUL!")
        print(f"     Contact: {status.get('contact_name', 'N/A')}")
        return True
    else:
        print(f"  ❌ Login may have failed: {status}")
        await page.screenshot(path="debug_login_failed.png")
        return False


# ═══════════════════════════════════════════════════════════════════
# API-BASED SCRAPING
# ═══════════════════════════════════════════════════════════════════

async def fetch_category_products(page, sub_category, is_food=True):
    """
    Fetch ALL products for a given category using the internal API.
    Uses page_size=200 for maximum efficiency.
    Returns list of product dicts.
    """
    all_products = []
    current_page = 1
    total_count = None

    while True:
        result = await page.evaluate(
            """
            async ([subCategory, isFood, pageNum, pageSize, postZip]) => {
                try {
                    const resp = await fetch(
                        `/api/ERP/search/productSearch?page=${pageNum}&page_size=${pageSize}&post_zip=${postZip}&ordering=relevance`,
                        {
                            method: 'POST',
                            headers: { 'Content-Type': 'application/json' },
                            body: JSON.stringify({
                                category: null,
                                sub_category: subCategory,
                                brand: null,
                                only_haggle_products: null,
                                food: isFood,
                                filters: { storage: ["DRY"] }
                            }),
                        }
                    );
                    const data = await resp.json();
                    return { 
                        status: resp.status, 
                        count: data.count, 
                        results: data.results || [],
                        next: data.next,
                    };
                } catch (err) {
                    return { error: err.message };
                }
            }
            """,
            [sub_category, is_food, current_page, API_PAGE_SIZE, POST_ZIP]
        )

        if "error" in result:
            print(f"      ✗ API error on page {current_page}: {result['error']}")
            break

        if result["status"] != 200:
            print(f"      ✗ HTTP {result['status']} on page {current_page}")
            break

        products = result.get("results", [])
        if total_count is None:
            total_count = result.get("count", 0)

        if not products:
            break

        all_products.extend(products)

        total_pages = math.ceil(total_count / API_PAGE_SIZE) if total_count else 1
        print(f"      Page {current_page}/{total_pages} — {len(products)} items (cumulative: {len(all_products)}/{total_count})")

        # Check if there are more pages
        if not result.get("next") or len(all_products) >= total_count:
            break

        current_page += 1
        await page.wait_for_timeout(1000)  # Polite delay between API calls

    return all_products, total_count or 0


async def scrape_all_categories(page):
    """Scrape all categories and return combined product list."""
    all_products = []
    category_stats = {}

    print("\n" + "=" * 60)
    print("📦 SCRAPING ALL CATEGORIES (DRY STORAGE)")
    print("=" * 60)

    # Combine food and non-food categories
    all_categories = [(cat, True) for cat in FOOD_CATEGORIES] + [(cat, False) for cat in NONFOOD_CATEGORIES]

    for i, (category, is_food) in enumerate(all_categories, 1):
        food_label = "Food" if is_food else "Non-Food"
        print(f"\n  [{i}/{len(all_categories)}] {category} ({food_label})")

        try:
            products, total = await fetch_category_products(page, category, is_food)

            if total == 0:
                print(f"      ○ No DRY products — skipping")
                continue

            # Process products
            for prod in products:
                processed = {
                    "category": category,
                    "food_type": food_label,
                    "manufacturer": clean_manufacturer(prod.get("brand_name", "")),
                    "product": prod.get("name", ""),
                    "description": prod.get("description_short", ""),
                    "upc": prod.get("upc", ""),
                    "delivered_price": prod.get("delivered_price", ""),
                    "delivered_case_price": prod.get("delivered_case_price", ""),
                    "price_per_unit": prod.get("per_unit_delivered_price", ""),
                    "price_per_oz": prod.get("per_oz_delivered_price", ""),
                    "pack_size_raw": prod.get("pack_size", ""),
                    "cases_per_pallet": prod.get("case_per_pallet", ""),
                    "lead_time_days": prod.get("lead_time_days", ""),
                    "min_pallet_qty": prod.get("min_pallet_quantity", ""),
                    "mixed_pallet": "Yes" if prod.get("for_mixed_pallet") else "No",
                    "available": "Yes" if prod.get("is_available") else "No",
                    "has_promo": "Yes" if prod.get("has_promo") else "No",
                    "main_category": prod.get("main_category", ""),
                    "sub_category": prod.get("sub_category", ""),
                    "product_id": prod.get("id", ""),
                    "product_url": f"https://epallet.com/product/{prod.get('slug', '')}",
                }
                all_products.append(processed)

            category_stats[category] = len(products)
            print(f"      ✅ {len(products)} products scraped")

            # Save checkpoint
            save_checkpoint(all_products)

        except Exception as e:
            print(f"      ❌ Error: {e}")
            category_stats[category] = 0

    return all_products, category_stats


# ═══════════════════════════════════════════════════════════════════
# DATA PROCESSING
# ═══════════════════════════════════════════════════════════════════

def clean_manufacturer(brand_name):
    """Clean brand name: remove '-EP' suffix."""
    if not brand_name:
        return ""
    cleaned = re.sub(r"-EP$", "", brand_name).strip()
    return cleaned


def parse_pack_size(raw):
    """
    Split pack size like '10/5 oz' → ('10', '5 oz').
    Handles: '12/21 oz', '4/12/1 oz', '6/28 oz', '24/0.97 oz', etc.
    """
    if not raw:
        return "", ""
    raw = raw.strip()
    match = re.match(r"^(\d+)/(.+)$", raw)
    if match:
        return match.group(1), match.group(2).strip()
    return "", raw


def safe_float(val):
    """Convert string to float safely."""
    if val is None or val == "":
        return None
    try:
        return float(str(val).replace(",", ""))
    except (ValueError, TypeError):
        return None


def save_checkpoint(products):
    """Save progress to CSV."""
    if not products:
        return
    df = pd.DataFrame(products)
    df.to_csv(CHECKPOINT_CSV, index=False, encoding="utf-8")


# ═══════════════════════════════════════════════════════════════════
# EXCEL EXPORT
# ═══════════════════════════════════════════════════════════════════

def export_to_excel(products, category_stats, filename):
    """Export to a beautifully formatted Excel workbook."""
    print("\n" + "=" * 60)
    print("📊 EXPORTING TO EXCEL")
    print("=" * 60)

    if not products:
        print("  ❌ No products to export!")
        return

    wb = Workbook()

    # ── Main Data Sheet ───────────────────────────────────────────
    ws = wb.active
    ws.title = "ePallet Products"

    headers = [
        ("Category", 22),
        ("Manufacturer", 25),
        ("Product", 45),
        ("Description", 35),
        ("UPC", 15),
        ("Delivered Price", 16),
        ("Case Price", 12),
        ("Price Per Unit", 14),
        ("Price Per Oz", 12),
        ("Pack Size", 14),
        ("Pack Count", 12),
        ("Unit Size", 12),
        ("Cases/Pallet", 13),
        ("Lead Time (Days)", 16),
        ("Min Pallet Qty", 14),
        ("Mixed Pallet", 13),
        ("Available", 10),
        ("Promo", 8),
        ("Main Category", 30),
        ("Sub Category", 30),
        ("Product URL", 20),
    ]

    # Styles
    header_fill = PatternFill(start_color="1B3A5C", end_color="1B3A5C", fill_type="solid")
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    alt_fill = PatternFill(start_color="F2F6FA", end_color="F2F6FA", fill_type="solid")
    currency_fmt = '$#,##0.00'
    thin_border = Border(
        left=Side(style="thin", color="D0D0D0"),
        right=Side(style="thin", color="D0D0D0"),
        top=Side(style="thin", color="D0D0D0"),
        bottom=Side(style="thin", color="D0D0D0"),
    )

    # Write headers
    for col_idx, (header_name, width) in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header_name)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = thin_border
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    ws.freeze_panes = "A2"
    ws.row_dimensions[1].height = 30

    # Write data
    currency_cols = {6, 7, 8, 9}  # Delivered, Case, Per Unit, Per Oz
    center_cols = {11, 13, 14, 15, 16, 17, 18}

    for row_idx, product in enumerate(products, 2):
        pack_count, unit_size = parse_pack_size(product.get("pack_size_raw", ""))

        row_data = [
            product.get("category", ""),
            product.get("manufacturer", ""),
            product.get("product", ""),
            product.get("description", ""),
            product.get("upc", ""),
            safe_float(product.get("delivered_price")),
            safe_float(product.get("delivered_case_price")),
            safe_float(product.get("price_per_unit")),
            safe_float(product.get("price_per_oz")),
            product.get("pack_size_raw", ""),
            pack_count,
            unit_size,
            product.get("cases_per_pallet", ""),
            product.get("lead_time_days", ""),
            product.get("min_pallet_qty", ""),
            product.get("mixed_pallet", ""),
            product.get("available", ""),
            product.get("has_promo", ""),
            product.get("main_category", ""),
            product.get("sub_category", ""),
            product.get("product_url", ""),
        ]

        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = thin_border
            cell.alignment = Alignment(vertical="center")

            if col_idx in currency_cols and value is not None:
                cell.number_format = currency_fmt
            if col_idx in center_cols:
                cell.alignment = Alignment(horizontal="center", vertical="center")

        # Alternate row shading
        if row_idx % 2 == 0:
            for col_idx in range(1, len(headers) + 1):
                ws.cell(row=row_idx, column=col_idx).fill = alt_fill

    # Auto filter
    last_col = get_column_letter(len(headers))
    ws.auto_filter.ref = f"A1:{last_col}{len(products) + 1}"

    # ── Summary Sheet ─────────────────────────────────────────────
    ws_sum = wb.create_sheet("Summary", 0)  # Insert at beginning
    ws_sum.sheet_properties.tabColor = "1B3A5C"

    title_font = Font(name="Calibri", size=16, bold=True, color="1B3A5C")
    subtitle_font = Font(name="Calibri", size=11, color="555555")
    section_font = Font(name="Calibri", size=13, bold=True, color="1B3A5C")

    ws_sum.cell(row=1, column=1, value="ePallet Product Scrape Report").font = title_font
    ws_sum.cell(row=2, column=1, value=f"Generated: {datetime.now().strftime('%B %d, %Y at %I:%M %p')}").font = subtitle_font
    ws_sum.cell(row=3, column=1, value=f"Filter: {STORAGE_FILTER} storage").font = subtitle_font
    ws_sum.cell(row=4, column=1, value=f"Account: {EMAIL}").font = subtitle_font

    ws_sum.cell(row=6, column=1, value="Overview").font = section_font
    ws_sum.cell(row=7, column=1, value="Total Products:")
    ws_sum.cell(row=7, column=2, value=len(products))
    ws_sum.cell(row=8, column=1, value="Categories with Products:")
    ws_sum.cell(row=8, column=2, value=sum(1 for v in category_stats.values() if v > 0))

    # Category breakdown
    ws_sum.cell(row=10, column=1, value="Category Breakdown").font = section_font

    cat_headers_list = ["Category", "Product Count", "% of Total"]
    for col_idx, h in enumerate(cat_headers_list, 1):
        cell = ws_sum.cell(row=11, column=col_idx, value=h)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="E8EEF4", end_color="E8EEF4", fill_type="solid")

    row = 12
    for cat, count in sorted(category_stats.items(), key=lambda x: -x[1]):
        if count > 0:
            ws_sum.cell(row=row, column=1, value=cat)
            ws_sum.cell(row=row, column=2, value=count)
            pct = count / len(products) * 100 if products else 0
            ws_sum.cell(row=row, column=3, value=f"{pct:.1f}%")
            row += 1

    ws_sum.column_dimensions["A"].width = 35
    ws_sum.column_dimensions["B"].width = 15
    ws_sum.column_dimensions["C"].width = 12

    # Save
    wb.save(filename)
    print(f"\n  ✅ Saved {len(products)} products to: {filename}")
    print(f"  📊 Categories with products: {sum(1 for v in category_stats.values() if v > 0)}")
    for cat, count in sorted(category_stats.items(), key=lambda x: -x[1]):
        if count > 0:
            print(f"     • {cat}: {count:,} items")


# ═══════════════════════════════════════════════════════════════════
# MAIN
# ═══════════════════════════════════════════════════════════════════

async def main():
    print("╔══════════════════════════════════════════════════════════════╗")
    print("║         ePallet.com Product Scraper v2.0 (API Mode)        ║")
    print("║         Target: All DRY storage products                   ║")
    print("╚══════════════════════════════════════════════════════════════╝")

    start_time = time.time()
    print(f"\n  Start: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    print(f"  Output: {OUTPUT_EXCEL}")
    print(f"  API page size: {API_PAGE_SIZE} (max efficiency)")
    print(f"  Categories to scan: {len(FOOD_CATEGORIES) + len(NONFOOD_CATEGORIES)}")

    async with async_playwright() as p:
        browser = await p.chromium.launch(
            headless=HEADLESS,
            args=["--disable-blink-features=AutomationControlled"],
        )
        context = await browser.new_context(
            viewport={"width": 1440, "height": 900},
            user_agent=(
                "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) "
                "AppleWebKit/537.36 (KHTML, like Gecko) "
                "Chrome/120.0.0.0 Safari/537.36"
            ),
        )
        page = await context.new_page()
        page.set_default_timeout(TIMEOUT)

        # Phase 1: Login
        login_ok = await login(page)
        if not login_ok:
            print("\n❌ Login failed. Exiting.")
            await browser.close()
            return

        # Phase 2: Scrape all categories via API
        all_products, category_stats = await scrape_all_categories(page)

        await browser.close()

    # Phase 3: Export to Excel
    if all_products:
        export_to_excel(all_products, category_stats, OUTPUT_EXCEL)

        # Also save CSV backup
        df = pd.DataFrame(all_products)
        csv_file = OUTPUT_EXCEL.replace(".xlsx", ".csv")
        df.to_csv(csv_file, index=False, encoding="utf-8")
        print(f"  📄 CSV backup: {csv_file}")

    # Final stats
    elapsed = time.time() - start_time
    print(f"\n{'═' * 60}")
    print(f"  ✅ SCRAPE COMPLETE")
    print(f"  Total products: {len(all_products):,}")
    print(f"  Time elapsed: {elapsed:.1f} seconds ({elapsed/60:.1f} minutes)")
    print(f"  Output: {OUTPUT_EXCEL}")
    print(f"{'═' * 60}")


if __name__ == "__main__":
    asyncio.run(main())
