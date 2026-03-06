"""
ePallet Data Analyzer — Flask Backend
Serves product data, filtering, analytics, and Excel export
"""

import os
import io
import csv
import json
import math
import datetime
from functools import wraps
from flask import Flask, request, jsonify, send_file, session
from flask_cors import CORS
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers

app = Flask(__name__, static_folder="static", static_url_path="")
app.secret_key = "epallet-scraper4000-secret-key-change-in-prod"
app.config['SESSION_COOKIE_SAMESITE'] = 'Lax'
app.config['SESSION_COOKIE_SECURE'] = False  # Set True if HTTPS only
CORS(app, supports_credentials=True)

# ── Config ──────────────────────────────────────────────────────────────────
BASE_DIR = os.path.dirname(os.path.abspath(__file__))
PROJECT_DIR = os.path.join(BASE_DIR, "..")
CSV_PATH = os.path.join(PROJECT_DIR, "epallet_dry_products.csv")

# Vercel has a read-only filesystem; use /tmp for writable paths
IS_VERCEL = os.environ.get("VERCEL", "") == "1" or os.path.exists("/vercel")
if IS_VERCEL:
    UPLOAD_DIR = "/tmp/uploads"
    UPLOAD_LOG = "/tmp/upload_history.json"
else:
    UPLOAD_DIR = os.path.join(PROJECT_DIR, "uploads")
    UPLOAD_LOG = os.path.join(PROJECT_DIR, "upload_history.json")

USERS = {
    "admin": "epallet2026",
    "michael": "LAFoods2026",
}

# Ensure upload directory exists
try:
    os.makedirs(UPLOAD_DIR, exist_ok=True)
except OSError:
    pass

# ── Load data once at startup ──────────────────────────────────────────────
def load_data():
    df = pd.read_csv(CSV_PATH)
    # Clean numeric columns
    for col in ["delivered_price", "delivered_case_price", "price_per_unit", "price_per_oz",
                "cases_per_pallet", "lead_time_days", "min_pallet_qty"]:
        df[col] = pd.to_numeric(df[col], errors="coerce")
    
    # Parse pack_size_raw into count and unit_size
    def parse_pack(raw):
        if pd.isna(raw) or not isinstance(raw, str):
            return pd.Series({"pack_count": None, "unit_size": None, "unit_measure": None})
        parts = raw.strip().split("/")
        if len(parts) >= 2:
            try:
                pack_count = int(parts[0].strip())
            except ValueError:
                try:
                    pack_count = float(parts[0].strip())
                except ValueError:
                    pack_count = None
            unit_part = "/".join(parts[1:]).strip()
            # Extract numeric and unit
            import re
            m = re.match(r"([\d.]+)\s*(.*)", unit_part)
            if m:
                try:
                    unit_size = float(m.group(1))
                except ValueError:
                    unit_size = None
                unit_measure = m.group(2).strip() if m.group(2) else ""
            else:
                unit_size = None
                unit_measure = unit_part
            return pd.Series({"pack_count": pack_count, "unit_size": unit_size, "unit_measure": unit_measure})
        return pd.Series({"pack_count": None, "unit_size": None, "unit_measure": None})
    
    parsed = df["pack_size_raw"].apply(parse_pack)
    df = pd.concat([df, parsed], axis=1)
    
    # Compute total weight/volume
    df["total_weight"] = df["pack_count"] * df["unit_size"]
    
    # Fill NaN for JSON serialization
    df = df.fillna("")
    
    return df

# ── Load data with error handling for serverless ───────────────────────────
DF = None
try:
    print(f"Loading data from: {CSV_PATH}")
    print(f"CSV exists: {os.path.exists(CSV_PATH)}")
    DF = load_data()
    print(f"Loaded {len(DF)} products across {DF['category'].nunique()} categories from {DF['manufacturer'].nunique()} manufacturers")
except Exception as e:
    print(f"ERROR loading data: {e}")
    import traceback
    traceback.print_exc()
    DF = pd.DataFrame()

# ── Upload history helper ──────────────────────────────────────────────────
def load_upload_history():
    if os.path.exists(UPLOAD_LOG):
        try:
            with open(UPLOAD_LOG, "r") as f:
                return json.load(f)
        except:
            return []
    return []

def save_upload_history(history):
    with open(UPLOAD_LOG, "w") as f:
        json.dump(history, f, indent=2)

# ── Auth decorator ─────────────────────────────────────────────────────────
def login_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        if not session.get("logged_in"):
            return jsonify({"error": "Unauthorized"}), 401
        return f(*args, **kwargs)
    return decorated

# ── Auth routes ────────────────────────────────────────────────────────────
@app.route("/api/login", methods=["POST"])
def login():
    data = request.json or {}
    username = data.get("username", "").strip().lower()
    password = data.get("password", "")
    if username in USERS and USERS[username] == password:
        session["logged_in"] = True
        session["username"] = username
        return jsonify({"ok": True, "username": username})
    return jsonify({"error": "Invalid credentials"}), 401

@app.route("/api/logout", methods=["POST"])
def logout():
    session.clear()
    return jsonify({"ok": True})

@app.route("/api/me")
def me():
    if session.get("logged_in"):
        return jsonify({"logged_in": True, "username": session.get("username")})
    return jsonify({"logged_in": False}), 401

# ── Data routes ────────────────────────────────────────────────────────────
@app.route("/api/summary")
@login_required
def summary():
    """Dashboard summary stats"""
    df = DF
    valid_prices = df[df["delivered_price"] != ""]["delivered_price"]
    valid_unit = df[df["price_per_unit"] != ""]["price_per_unit"]
    
    return jsonify({
        "total_products": len(df),
        "total_categories": int(df["category"].nunique()),
        "total_manufacturers": int(df["manufacturer"].nunique()),
        "avg_delivered_price": round(float(valid_prices.mean()), 2) if len(valid_prices) else 0,
        "avg_unit_price": round(float(valid_unit.mean()), 2) if len(valid_unit) else 0,
        "min_unit_price": round(float(valid_unit.min()), 2) if len(valid_unit) else 0,
        "max_unit_price": round(float(valid_unit.max()), 2) if len(valid_unit) else 0,
        "total_with_promo": int((df["has_promo"] == "Yes").sum()),
        "total_mixed_pallet": int((df["mixed_pallet"] == "Yes").sum()),
    })

@app.route("/api/filters")
@login_required
def filters():
    """Get all filter options"""
    df = DF
    categories = sorted(df["category"].unique().tolist())
    manufacturers = sorted(df["manufacturer"].unique().tolist())
    lead_times = sorted([x for x in df["lead_time_days"].unique().tolist() if x != ""])
    
    return jsonify({
        "categories": categories,
        "manufacturers": manufacturers,
        "lead_times": lead_times,
    })

def apply_filters(df, params):
    """Apply query params as filters to the dataframe"""
    # Category filter
    cats = params.get("category")
    if cats:
        cat_list = cats.split(",")
        df = df[df["category"].isin(cat_list)]
    
    # Manufacturer filter
    mfrs = params.get("manufacturer")
    if mfrs:
        mfr_list = mfrs.split(",")
        df = df[df["manufacturer"].isin(mfr_list)]
    
    # Search
    search = params.get("search", "").strip()
    if search:
        mask = (
            df["product"].str.contains(search, case=False, na=False) |
            df["manufacturer"].str.contains(search, case=False, na=False) |
            df["description"].str.contains(search, case=False, na=False) |
            df["upc"].astype(str).str.contains(search, case=False, na=False)
        )
        df = df[mask]
    
    # Price range
    min_price = params.get("min_price")
    max_price = params.get("max_price")
    if min_price:
        df = df[df["price_per_unit"] != ""]
        df = df[df["price_per_unit"].astype(float) >= float(min_price)]
    if max_price:
        df = df[df["price_per_unit"] != ""]
        df = df[df["price_per_unit"].astype(float) <= float(max_price)]
    
    # Promo only
    if params.get("promo_only") == "true":
        df = df[df["has_promo"] == "Yes"]
    
    # Mixed pallet only
    if params.get("mixed_pallet_only") == "true":
        df = df[df["mixed_pallet"] == "Yes"]
    
    # Available only
    if params.get("available_only") == "true":
        df = df[df["available"] == "Yes"]
    
    return df

@app.route("/api/products")
@login_required
def products():
    """Paginated, filtered, sorted product list"""
    df = apply_filters(DF.copy(), request.args)
    
    # Sorting
    sort_by = request.args.get("sort_by", "product")
    sort_dir = request.args.get("sort_dir", "asc")
    
    if sort_by in df.columns:
        ascending = sort_dir == "asc"
        # For numeric columns, ensure proper sorting
        if sort_by in ["delivered_price", "delivered_case_price", "price_per_unit", 
                        "price_per_oz", "cases_per_pallet", "lead_time_days", "total_weight"]:
            df[sort_by] = pd.to_numeric(df[sort_by], errors="coerce")
        df = df.sort_values(by=sort_by, ascending=ascending, na_position="last")
    
    total = len(df)
    
    # Pagination
    page = int(request.args.get("page", 1))
    per_page = int(request.args.get("per_page", 50))
    start = (page - 1) * per_page
    end = start + per_page
    
    page_df = df.iloc[start:end]
    
    # Convert to records, replacing NaN
    records = page_df.fillna("").to_dict(orient="records")
    
    return jsonify({
        "products": records,
        "total": total,
        "page": page,
        "per_page": per_page,
        "total_pages": math.ceil(total / per_page),
    })

# ── New Analytics routes ───────────────────────────────────────────────────

@app.route("/api/analytics/best-value")
@login_required
def best_value():
    """Best value products — lowest price-per-oz with good availability"""
    df = apply_filters(DF.copy(), request.args)
    df_valid = df[(df["price_per_oz"] != "") & (df["cases_per_pallet"] != "")].copy()
    df_valid["price_per_oz"] = pd.to_numeric(df_valid["price_per_oz"], errors="coerce")
    df_valid["cases_per_pallet"] = pd.to_numeric(df_valid["cases_per_pallet"], errors="coerce")
    df_valid["price_per_unit"] = pd.to_numeric(df_valid["price_per_unit"], errors="coerce")
    df_valid["delivered_price"] = pd.to_numeric(df_valid["delivered_price"], errors="coerce")
    df_valid = df_valid.dropna(subset=["price_per_oz", "cases_per_pallet"])
    
    # Value score: lower price_per_oz + higher cases_per_pallet = better value
    if len(df_valid) > 0:
        poz_norm = 1 - (df_valid["price_per_oz"] - df_valid["price_per_oz"].min()) / (df_valid["price_per_oz"].max() - df_valid["price_per_oz"].min() + 0.0001)
        cpp_norm = (df_valid["cases_per_pallet"] - df_valid["cases_per_pallet"].min()) / (df_valid["cases_per_pallet"].max() - df_valid["cases_per_pallet"].min() + 0.0001)
        df_valid["value_score"] = (poz_norm * 0.6 + cpp_norm * 0.4).round(3)
    else:
        df_valid["value_score"] = 0
    
    limit = int(request.args.get("limit", 30))
    direction = request.args.get("direction", "desc")
    ascending = direction == "asc"
    df_sorted = df_valid.sort_values("value_score", ascending=ascending).head(limit)
    
    cols = ["product", "manufacturer", "category", "price_per_unit", "price_per_oz",
            "delivered_price", "pack_size_raw", "cases_per_pallet", "lead_time_days",
            "available", "mixed_pallet", "value_score"]
    return jsonify(df_sorted[cols].fillna("").to_dict(orient="records"))

@app.route("/api/analytics/pallet-efficiency")
@login_required
def pallet_efficiency():
    """Pallet efficiency — cost per pallet vs cases per pallet"""
    df = apply_filters(DF.copy(), request.args)
    df_valid = df[(df["delivered_price"] != "") & (df["cases_per_pallet"] != "")].copy()
    df_valid["delivered_price"] = pd.to_numeric(df_valid["delivered_price"], errors="coerce")
    df_valid["cases_per_pallet"] = pd.to_numeric(df_valid["cases_per_pallet"], errors="coerce")
    df_valid = df_valid.dropna(subset=["delivered_price", "cases_per_pallet"])
    df_valid["cost_per_case"] = (df_valid["delivered_price"] / df_valid["cases_per_pallet"]).round(2)
    
    grouped = df_valid.groupby("category").agg(
        avg_cases_per_pallet=("cases_per_pallet", "mean"),
        avg_delivered=("delivered_price", "mean"),
        avg_cost_per_case=("cost_per_case", "mean"),
        product_count=("product", "count"),
    ).round(2).reset_index()
    
    return jsonify(grouped.sort_values("avg_cost_per_case").to_dict(orient="records"))

@app.route("/api/analytics/cost-per-oz")
@login_required
def cost_per_oz_analysis():
    """Cost per oz breakdown by category"""
    df = apply_filters(DF.copy(), request.args)
    df_valid = df[df["price_per_oz"] != ""].copy()
    df_valid["price_per_oz"] = pd.to_numeric(df_valid["price_per_oz"], errors="coerce")
    df_valid = df_valid.dropna(subset=["price_per_oz"])
    
    grouped = df_valid.groupby("category").agg(
        avg_per_oz=("price_per_oz", "mean"),
        min_per_oz=("price_per_oz", "min"),
        max_per_oz=("price_per_oz", "max"),
        median_per_oz=("price_per_oz", "median"),
        count=("product", "count"),
    ).round(4).reset_index()
    
    return jsonify(grouped.sort_values("avg_per_oz").to_dict(orient="records"))

@app.route("/api/analytics/availability")
@login_required
def availability_analysis():
    """Availability and mixed pallet analysis"""
    df = apply_filters(DF.copy(), request.args)
    
    total = len(df)
    available_yes = int((df["available"] == "Yes").sum())
    available_no = total - available_yes
    mixed_yes = int((df["mixed_pallet"] == "Yes").sum())
    mixed_no = total - mixed_yes
    promo_yes = int((df["has_promo"] == "Yes").sum())
    
    # By category
    cat_avail = df.groupby("category").agg(
        total=("product", "count"),
        available=("available", lambda x: (x == "Yes").sum()),
        mixed_pallet=("mixed_pallet", lambda x: (x == "Yes").sum()),
    ).reset_index()
    cat_avail["avail_pct"] = (cat_avail["available"] / cat_avail["total"] * 100).round(1)
    cat_avail["mixed_pct"] = (cat_avail["mixed_pallet"] / cat_avail["total"] * 100).round(1)
    
    return jsonify({
        "totals": {
            "total": total,
            "available": available_yes,
            "unavailable": available_no,
            "mixed_pallet": mixed_yes,
            "non_mixed": mixed_no,
            "promo": promo_yes,
        },
        "by_category": cat_avail.to_dict(orient="records"),
    })

@app.route("/api/analytics/lead-time-summary")
@login_required
def lead_time_summary():
    """Lead time stats by category"""
    df = apply_filters(DF.copy(), request.args)
    df_valid = df[df["lead_time_days"] != ""].copy()
    df_valid["lead_time_days"] = pd.to_numeric(df_valid["lead_time_days"], errors="coerce")
    df_valid = df_valid.dropna(subset=["lead_time_days"])
    
    grouped = df_valid.groupby("category").agg(
        avg_lead=("lead_time_days", "mean"),
        min_lead=("lead_time_days", "min"),
        max_lead=("lead_time_days", "max"),
        count=("product", "count"),
    ).round(1).reset_index()
    
    return jsonify(grouped.sort_values("avg_lead").to_dict(orient="records"))

# ── Existing analytics routes ─────────────────────────────────────────────
@app.route("/api/analytics/by-category")
@login_required
def analytics_by_category():
    """Price analytics grouped by category"""
    df = apply_filters(DF.copy(), request.args)
    df_valid = df[df["price_per_unit"] != ""].copy()
    df_valid["price_per_unit"] = df_valid["price_per_unit"].astype(float)
    df_valid["delivered_price"] = pd.to_numeric(df_valid["delivered_price"], errors="coerce")
    
    grouped = df_valid.groupby("category").agg(
        count=("product", "count"),
        avg_unit_price=("price_per_unit", "mean"),
        min_unit_price=("price_per_unit", "min"),
        max_unit_price=("price_per_unit", "max"),
        median_unit_price=("price_per_unit", "median"),
        avg_delivered=("delivered_price", "mean"),
        total_value=("delivered_price", "sum"),
    ).round(2).reset_index()
    
    return jsonify(grouped.to_dict(orient="records"))

@app.route("/api/analytics/by-manufacturer")
@login_required
def analytics_by_manufacturer():
    """Price analytics grouped by manufacturer"""
    df = apply_filters(DF.copy(), request.args)
    df_valid = df[df["price_per_unit"] != ""].copy()
    df_valid["price_per_unit"] = df_valid["price_per_unit"].astype(float)
    df_valid["delivered_price"] = pd.to_numeric(df_valid["delivered_price"], errors="coerce")
    
    grouped = df_valid.groupby("manufacturer").agg(
        count=("product", "count"),
        avg_unit_price=("price_per_unit", "mean"),
        min_unit_price=("price_per_unit", "min"),
        max_unit_price=("price_per_unit", "max"),
        avg_delivered=("delivered_price", "mean"),
        categories=("category", "nunique"),
    ).round(2).reset_index()
    
    # Sort by count descending, take top N
    limit = int(request.args.get("limit", 30))
    sort = request.args.get("sort", "count")
    grouped = grouped.sort_values(by=sort, ascending=False).head(limit)
    
    return jsonify(grouped.to_dict(orient="records"))

@app.route("/api/analytics/price-distribution")
@login_required
def price_distribution():
    """Price distribution histogram data"""
    df = apply_filters(DF.copy(), request.args)
    df_valid = df[df["price_per_unit"] != ""].copy()
    df_valid["price_per_unit"] = df_valid["price_per_unit"].astype(float)
    
    # Remove extreme outliers for better visualization
    q99 = df_valid["price_per_unit"].quantile(0.99)
    df_clipped = df_valid[df_valid["price_per_unit"] <= q99]
    
    bins = int(request.args.get("bins", 30))
    counts, edges = pd.cut(df_clipped["price_per_unit"], bins=bins, retbins=True)
    hist = df_clipped.groupby(counts, observed=True).size().reset_index(name="count")
    
    result = []
    for i, row in hist.iterrows():
        interval = row.iloc[0]
        result.append({
            "bin_start": round(interval.left, 2),
            "bin_end": round(interval.right, 2),
            "label": f"${interval.left:.2f}-${interval.right:.2f}",
            "count": int(row["count"]),
        })
    
    return jsonify(result)

@app.route("/api/analytics/price-vs-volume")
@login_required
def price_vs_volume():
    """Scatter plot data: unit price vs total weight, by category"""
    df = apply_filters(DF.copy(), request.args)
    df_valid = df[(df["price_per_unit"] != "") & (df["total_weight"] != "")].copy()
    df_valid["price_per_unit"] = df_valid["price_per_unit"].astype(float)
    df_valid["total_weight"] = df_valid["total_weight"].astype(float)
    
    # Sample if too many points
    if len(df_valid) > 2000:
        df_valid = df_valid.sample(2000, random_state=42)
    
    records = df_valid[["product", "manufacturer", "category", "price_per_unit", 
                         "total_weight", "unit_measure", "pack_size_raw"]].to_dict(orient="records")
    return jsonify(records)

@app.route("/api/analytics/lead-time")
@login_required
def lead_time_analysis():
    """Lead time distribution by category"""
    df = apply_filters(DF.copy(), request.args)
    df_valid = df[df["lead_time_days"] != ""].copy()
    df_valid["lead_time_days"] = df_valid["lead_time_days"].astype(int)
    
    grouped = df_valid.groupby(["category", "lead_time_days"]).size().reset_index(name="count")
    return jsonify(grouped.to_dict(orient="records"))

@app.route("/api/analytics/top-value")
@login_required
def top_value_products():
    """Top/Bottom products by various metrics"""
    df = apply_filters(DF.copy(), request.args)
    metric = request.args.get("metric", "price_per_unit")
    direction = request.args.get("direction", "desc")
    limit = int(request.args.get("limit", 20))
    
    df_valid = df[df[metric] != ""].copy()
    df_valid[metric] = pd.to_numeric(df_valid[metric], errors="coerce")
    df_valid = df_valid.dropna(subset=[metric])
    
    ascending = direction == "asc"
    df_sorted = df_valid.sort_values(by=metric, ascending=ascending).head(limit)
    
    cols = ["product", "manufacturer", "category", "delivered_price", "price_per_unit",
            "price_per_oz", "pack_size_raw", "cases_per_pallet", "total_weight"]
    return jsonify(df_sorted[cols].fillna("").to_dict(orient="records"))

@app.route("/api/analytics/category-manufacturer-matrix")
@login_required  
def category_manufacturer_matrix():
    """Heatmap data: categories vs top manufacturers"""
    df = apply_filters(DF.copy(), request.args)
    
    # Get top N manufacturers by product count
    limit = int(request.args.get("limit", 15))
    top_mfrs = df["manufacturer"].value_counts().head(limit).index.tolist()
    
    df_filtered = df[df["manufacturer"].isin(top_mfrs)]
    
    matrix = df_filtered.groupby(["category", "manufacturer"]).agg(
        count=("product", "count"),
        avg_price=("price_per_unit", lambda x: round(pd.to_numeric(x, errors="coerce").mean(), 2))
    ).reset_index()
    
    return jsonify({
        "data": matrix.fillna(0).to_dict(orient="records"),
        "categories": sorted(df_filtered["category"].unique().tolist()),
        "manufacturers": top_mfrs,
    })

@app.route("/api/analytics/compare-products")
@login_required
def compare_products():
    """Compare specific products side by side"""
    ids = request.args.get("product_ids", "")
    if not ids:
        return jsonify([])
    
    id_list = [int(x) for x in ids.split(",")]
    df_match = DF[DF["product_id"].isin(id_list)]
    return jsonify(df_match.fillna("").to_dict(orient="records"))

# ── Export routes ──────────────────────────────────────────────────────────
@app.route("/api/export/excel")
@login_required
def export_excel():
    """Export filtered data as formatted Excel"""
    df = apply_filters(DF.copy(), request.args)
    
    # Sorting
    sort_by = request.args.get("sort_by", "category")
    sort_dir = request.args.get("sort_dir", "asc")
    if sort_by in df.columns:
        df = df.sort_values(by=sort_by, ascending=(sort_dir == "asc"), na_position="last")
    
    wb = Workbook()
    ws = wb.active
    ws.title = "ePallet Products"
    
    # Styles
    header_font = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="1B3A5C", end_color="1B3A5C", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    currency_format = '#,##0.00'
    border = Border(
        left=Side(style="thin", color="D0D0D0"),
        right=Side(style="thin", color="D0D0D0"),
        top=Side(style="thin", color="D0D0D0"),
        bottom=Side(style="thin", color="D0D0D0"),
    )
    alt_fill = PatternFill(start_color="F2F6FA", end_color="F2F6FA", fill_type="solid")
    
    columns = [
        ("Category", "category", 22),
        ("Manufacturer", "manufacturer", 25),
        ("Product", "product", 45),
        ("Description", "description", 35),
        ("UPC", "upc", 16),
        ("Delivered Price", "delivered_price", 16),
        ("Case Price", "delivered_case_price", 14),
        ("Price/Unit", "price_per_unit", 12),
        ("Price/Oz", "price_per_oz", 11),
        ("Pack Size", "pack_size_raw", 14),
        ("Pack Count", "pack_count", 12),
        ("Unit Size", "unit_size", 11),
        ("Unit", "unit_measure", 8),
        ("Total Weight", "total_weight", 13),
        ("Cases/Pallet", "cases_per_pallet", 13),
        ("Lead Time (days)", "lead_time_days", 15),
        ("Mixed Pallet", "mixed_pallet", 13),
        ("Available", "available", 10),
        ("Promo", "has_promo", 8),
    ]
    
    # Write headers
    for col_idx, (header, _, width) in enumerate(columns, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = border
        ws.column_dimensions[cell.column_letter].width = width
    
    # Write data
    currency_cols = {"delivered_price", "delivered_case_price", "price_per_unit", "price_per_oz"}
    number_cols = {"pack_count", "unit_size", "total_weight", "cases_per_pallet", "lead_time_days"}
    
    for row_idx, (_, row) in enumerate(df.iterrows(), 2):
        for col_idx, (_, field, _) in enumerate(columns, 1):
            value = row.get(field, "")
            cell = ws.cell(row=row_idx, column=col_idx)
            
            if field in currency_cols and value != "":
                try:
                    cell.value = float(value)
                    cell.number_format = '$#,##0.00'
                except (ValueError, TypeError):
                    cell.value = value
            elif field in number_cols and value != "":
                try:
                    cell.value = float(value)
                    cell.number_format = '#,##0.##'
                except (ValueError, TypeError):
                    cell.value = value
            else:
                cell.value = str(value) if value != "" else ""
            
            cell.border = border
            cell.alignment = Alignment(vertical="center")
            if row_idx % 2 == 0:
                cell.fill = alt_fill
    
    # Freeze header row
    ws.freeze_panes = "A2"
    
    # Auto-filter
    ws.auto_filter.ref = ws.dimensions
    
    # Summary sheet
    ws2 = wb.create_sheet("Summary")
    ws2.cell(row=1, column=1, value="ePallet Data Summary").font = Font(bold=True, size=14)
    ws2.cell(row=2, column=1, value=f"Generated: March 6, 2026")
    ws2.cell(row=3, column=1, value=f"Total Products: {len(df)}")
    ws2.cell(row=4, column=1, value=f"Categories: {df['category'].nunique()}")
    ws2.cell(row=5, column=1, value=f"Manufacturers: {df['manufacturer'].nunique()}")
    
    if len(df) > 0 and "price_per_unit" in df.columns:
        valid = pd.to_numeric(df["price_per_unit"], errors="coerce").dropna()
        if len(valid) > 0:
            ws2.cell(row=7, column=1, value="Price Per Unit Stats").font = Font(bold=True)
            ws2.cell(row=8, column=1, value=f"Average: ${valid.mean():.2f}")
            ws2.cell(row=9, column=1, value=f"Median: ${valid.median():.2f}")
            ws2.cell(row=10, column=1, value=f"Min: ${valid.min():.2f}")
            ws2.cell(row=11, column=1, value=f"Max: ${valid.max():.2f}")
    
    # Save to buffer
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    
    filename = "epallet_products"
    cats = request.args.get("category")
    if cats:
        filename += f"_{cats.replace(',', '_')}"
    filename += ".xlsx"
    
    return send_file(buf, download_name=filename, as_attachment=True,
                     mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

@app.route("/api/export/csv")
@login_required
def export_csv():
    """Export filtered data as CSV"""
    df = apply_filters(DF.copy(), request.args)
    buf = io.StringIO()
    df.to_csv(buf, index=False)
    
    return send_file(
        io.BytesIO(buf.getvalue().encode()),
        download_name="epallet_products.csv",
        as_attachment=True,
        mimetype="text/csv"
    )

# ── Upload route ───────────────────────────────────────────────────────────
@app.route("/api/upload", methods=["POST"])
@login_required
def upload_data():
    """Upload CSV or XLSX file to merge into the dataset"""
    global DF
    
    if "file" not in request.files:
        return jsonify({"error": "No file provided"}), 400
    
    file = request.files["file"]
    if not file.filename:
        return jsonify({"error": "No file selected"}), 400
    
    filename = file.filename.lower()
    if not (filename.endswith(".csv") or filename.endswith(".xlsx") or filename.endswith(".xls")):
        return jsonify({"error": "Only CSV and Excel files are supported"}), 400
    
    try:
        # Read the uploaded file
        if filename.endswith(".csv"):
            upload_df = pd.read_csv(file)
        else:
            upload_df = pd.read_excel(file)
        
        if len(upload_df) == 0:
            return jsonify({"error": "File is empty"}), 400
        
        # Normalize column names: lowercase, strip, replace spaces with underscores
        upload_df.columns = [c.strip().lower().replace(" ", "_").replace("/", "_per_") for c in upload_df.columns]
        
        # Column mapping — try to match common column names to our schema
        col_map = {
            "delivered_price": ["delivered_price", "delivered price", "pallet_price", "pallet price", "price"],
            "delivered_case_price": ["delivered_case_price", "case_price", "case price"],
            "price_per_unit": ["price_per_unit", "unit_price", "unit price", "price_per_unit"],
            "price_per_oz": ["price_per_oz", "price per oz", "oz_price"],
            "pack_size_raw": ["pack_size_raw", "pack_size", "pack size", "size"],
            "cases_per_pallet": ["cases_per_pallet", "cases per pallet", "cases_pallet", "cs_per_pallet"],
            "lead_time_days": ["lead_time_days", "lead_time", "lead time", "leadtime"],
            "category": ["category", "cat", "product_category"],
            "manufacturer": ["manufacturer", "brand", "mfr", "vendor", "supplier"],
            "product": ["product", "product_name", "name", "item", "description", "item_name"],
            "upc": ["upc", "upc_code", "barcode", "ean", "gtin"],
            "description": ["description", "desc", "product_description", "item_description"],
            "mixed_pallet": ["mixed_pallet", "mixed pallet", "mixed"],
            "available": ["available", "in_stock", "availability", "in stock"],
            "has_promo": ["has_promo", "promo", "promotion", "on_sale"],
            "min_pallet_qty": ["min_pallet_qty", "min_qty", "minimum_qty", "moq"],
            "food_type": ["food_type", "type"],
            "main_category": ["main_category"],
            "sub_category": ["sub_category"],
        }
        
        # Apply column mapping
        rename_map = {}
        for target, sources in col_map.items():
            for src in sources:
                normalized = src.strip().lower().replace(" ", "_").replace("/", "_per_")
                if normalized in upload_df.columns and target not in upload_df.columns:
                    rename_map[normalized] = target
                    break
        upload_df = upload_df.rename(columns=rename_map)
        
        # Determine which required columns we have
        our_cols = set(DF.columns)
        upload_cols = set(upload_df.columns)
        matched_cols = our_cols & upload_cols
        
        # Must have at least product or description to be useful
        if "product" not in upload_df.columns and "description" not in upload_df.columns:
            # Try to use first text column as product
            for c in upload_df.columns:
                if upload_df[c].dtype == "object" and c not in ["upc", "category", "manufacturer"]:
                    upload_df = upload_df.rename(columns={c: "product"})
                    break
        
        if "product" not in upload_df.columns:
            return jsonify({"error": "Could not identify a product/name column. Please ensure your file has a 'Product' column."}), 400
        
        # Fill missing columns with empty strings
        for col in DF.columns:
            if col not in upload_df.columns:
                upload_df[col] = ""
        
        # Keep only our columns, in the right order
        upload_df = upload_df[[c for c in DF.columns if c in upload_df.columns]]
        
        # Add missing columns as empty
        for c in DF.columns:
            if c not in upload_df.columns:
                upload_df[c] = ""
        upload_df = upload_df[DF.columns]
        
        # Tag with data source
        upload_df["data_source"] = "upload:" + file.filename
        upload_df["upload_date"] = datetime.datetime.now().strftime("%Y-%m-%d %H:%M")
        
        # Add data_source to main DF if not present
        if "data_source" not in DF.columns:
            DF["data_source"] = "epallet_scrape"
            DF["upload_date"] = "2026-03-05"
        
        # Clean numeric columns
        for col in ["delivered_price", "delivered_case_price", "price_per_unit", "price_per_oz",
                    "cases_per_pallet", "lead_time_days", "min_pallet_qty"]:
            upload_df[col] = pd.to_numeric(upload_df[col], errors="coerce")
        
        # Generate product IDs for new rows
        max_id = 0
        if "product_id" in DF.columns:
            existing_ids = pd.to_numeric(DF["product_id"], errors="coerce").dropna()
            if len(existing_ids) > 0:
                max_id = int(existing_ids.max())
        upload_df["product_id"] = range(max_id + 1, max_id + 1 + len(upload_df))
        
        rows_before = len(DF)
        
        # Merge: append new data
        DF = pd.concat([DF, upload_df], ignore_index=True)
        DF = DF.fillna("")
        
        rows_added = len(upload_df)
        rows_after = len(DF)
        
        # Save updated CSV
        DF.to_csv(CSV_PATH, index=False)
        
        # Save the uploaded file
        save_path = os.path.join(UPLOAD_DIR, f"{datetime.datetime.now().strftime('%Y%m%d_%H%M%S')}_{file.filename}")
        # We need to reset the file pointer, but it's been read — save from df
        upload_df.to_csv(save_path.replace('.xlsx', '.csv').replace('.xls', '.csv'), index=False)
        
        # Log upload
        history = load_upload_history()
        history.append({
            "filename": file.filename,
            "timestamp": datetime.datetime.now().isoformat(),
            "rows_added": rows_added,
            "columns_matched": list(matched_cols),
            "columns_missing": list(our_cols - upload_cols),
            "uploaded_by": session.get("username", "unknown"),
            "total_after": rows_after,
        })
        save_upload_history(history)
        
        return jsonify({
            "ok": True,
            "rows_added": rows_added,
            "total_products": rows_after,
            "columns_matched": sorted(list(matched_cols)),
            "columns_missing": sorted(list(our_cols - upload_cols)),
            "filename": file.filename,
        })
    
    except Exception as e:
        return jsonify({"error": f"Failed to process file: {str(e)}"}), 500

@app.route("/api/upload/history")
@login_required
def upload_history():
    """Get upload history"""
    history = load_upload_history()
    return jsonify(history)

@app.route("/api/data-sources")
@login_required
def data_sources():
    """Get data source breakdown"""
    df = DF
    if "data_source" not in df.columns:
        return jsonify([{"source": "epallet_scrape", "count": len(df), "pct": 100.0}])
    
    sources = df["data_source"].value_counts().reset_index()
    sources.columns = ["source", "count"]
    sources["pct"] = (sources["count"] / sources["count"].sum() * 100).round(1)
    return jsonify(sources.to_dict(orient="records"))

# ── Serve frontend ─────────────────────────────────────────────────────────
@app.route("/")
def index():
    return send_file("static/index.html")

if __name__ == "__main__":
    app.run(debug=True, port=5050)
